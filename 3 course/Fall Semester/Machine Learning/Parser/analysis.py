import pandas as pd
import numpy as np
import json
from openpyxl import load_workbook
import re

def get_words_dict(path):
    wb = load_workbook(filename = path)
    
    # extract words dict for each category from excel file
    word_dict = {}

    for range in wb.sheetnames:
        ws = wb[range]
        words = []
        column = 'A'
        row_i = 2
        while True:
            word = ws[column + str(row_i)].value
            if (word and len(word) > 0):
                words.append(word.lower().strip())
                row_i = row_i + 1
            else: 
                break    
        word_dict[range.lower().strip()] = words

    return word_dict

def drop_inappropriate_tweets(df, words_dict):
    new_df = pd.DataFrame()
    for category in list(words_dict.keys()):
        category_df = df[df['category'] == category]
        regex = ''.join(['.*' + word + '.*|' for word in words_dict[category]])[:-1]
        
        len_before = len(category_df)
        category_df = category_df[category_df['content'].str.match(regex, case = False)]
        len_after = len(category_df)
        
        print(f'Number of elements in category "{category}": before - {len_before}, after - {len_after}')
        new_df = pd.concat([new_df, category_df])
    return new_df

def calculate_words_frequency(df, words_dict):
    count_dict = {}
    for category in list(words_dict.keys()):
        category_dict = {}
        category_df = df[df['category'] == category]
        for word in words_dict[category]:
            regex = '.*' + word + '.*'
            category_dict[word] = category_df['content'].str.count(regex)
        regex = ''.join(['.*' + word + '.*|' for word in words_dict[category]])[:-1]
        category_dict = sorted(category_dict.items(), key=lambda x:x[1])
        count_dict[category] = category_dict
        print(f'Calculate frequency of words in "{category}" category')
    return count_dict

data = None
with open("./data.json", "r") as f:
        data = f.read()
        data = data[:-1]
        data = "[" + data  + "]"

data = json.loads(data)
df1 = pd.DataFrame(data)
df2 = pd.read_csv('./processed_data.csv')

df = pd.concat([df1, df2], ignore_index = True)

df.drop_duplicates(subset='content', inplace=True)
df.dropna(subset='content', inplace=True)
df.drop(columns=['Unnamed: 0'], inplace=True)

words_dict = get_words_dict(r'./words/word_dicionaries.xlsx')

processed_df = drop_inappropriate_tweets(df, words_dict)

print(processed_df.info())
print(f'Number of tweets in each category:\n {processed_df["category"].value_counts()}')

processed_df.to_excel(r'final_dataset.xlsx', columns=['content', 'category'])