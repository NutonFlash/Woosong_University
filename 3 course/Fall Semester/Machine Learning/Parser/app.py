import asyncio
from twscrape import API
from twscrape.logger import set_log_level
import asyncio
import preprocessor as p
import json
from openpyxl import load_workbook
import datetime
from dateutil import relativedelta

def format_result(tweet, category):
    result = {
        "id": tweet.id,
        "content": p.clean(tweet.rawContent),
        "category": category
    }
    return json.dumps(result)

def words_search_query(keywords):
    max_len = 500
    reserve_len = 15
    dates_len = len('since:0000-00-00 until:0000-00-00')
    max_words_query_len = max_len - len('lang:en') - dates_len - reserve_len

    current_length = 0
    result_strings = []
    current_string = ""

    for word in keywords:
        # Check if adding the next word would exceed the maximum length
        if word.rfind(' ') > 0:
            word = '(' + word + ')'
        
        if current_length + len(word) + len(' OR ') > max_words_query_len:
            # If yes, start a new string
            result_strings.append(current_string[:-len(' OR ')])  # Remove the last ' OR '
            current_string = word + ' OR '
            current_length = len(current_string)
        else:
            # If not, add the word to the current string
            current_string += word + ' OR '
            current_length += len(word) + len(' OR ')

    # Add the last string to the result_strings
    result_strings.append(current_string[:-len(' OR ')])

    return result_strings

def get_words_dict(path):
    wb = load_workbook(filename = path)
    
    # extract words dict for each category from excel file
    word_dict = {}

    for range in wb.sheetnames:
        ws = wb[range]
        words = []
        column = 'A'
        row_i = 2
        while True:
            word = ws[column + str(row_i)].value
            if (word and len(word) > 0):
                words.append(word.lower().strip())
                row_i = row_i + 1
            else: 
                break    
        word_dict[range.lower().strip()] = words

    return word_dict

def get_dates_dict(start_date, end_date, month_step):
    dates = {}
    
    next_date = start_date + relativedelta.relativedelta(months=month_step)
    
    while next_date < end_date:
        dates[start_date.strftime('%Y-%m-%d')] = next_date.strftime('%Y-%m-%d')
        start_date = next_date
        next_date = next_date + relativedelta.relativedelta(months=month_step)

    return dates

def create_query_queue(options):
    queries = []
    categories = options['categories']
    for category in categories:
        word_queries = words_search_query(options['words'].get(category, []))
        for hashtag in options['hashtags'].get(category, []):
            for start_date, end_date in options['dates'].items():
                for word_query in word_queries:
                    queries.append({
                        'category': category,
                        'query': f'{word_query} {hashtag} lang:en since:{start_date} until:{end_date}'
                    })
                    
        for user in options['users'].get(category, []):
            for start_date, end_date in options['dates'].items():
                for word_query in word_queries:
                    queries.append({
                        'category': category,
                        'query': f'{word_query} from:{user} lang:en since:{start_date} until:{end_date}'
                    })
    return queries

async def worker(queue: asyncio.Queue, api: API):
    with open("data.json", "a", encoding='utf-8') as data_file:
        while True:
            query = await queue.get()
            try:
                print(f'Size of the queue: {queue.qsize()}\n')
                async for tweet in api.search(query['query'], limit=1):
                    data_file.write(format_result(tweet, query['category']) + ',')
            except Exception as e:
                print(f"Error on {query['query']} - {type(e)}")
            finally:
                queue.task_done()

async def main():
    api = API()

    # change log level, default info
    set_log_level("DEBUG")

    words_dict = get_words_dict(r'./words/word_dicionaries.xlsx')
    accaunts_dict = {
        'religion': list(dict.fromkeys(['JohnPiper', 'Fervr', 'Relevant', 'bethanyhamilton', 'wwmonitor', 'timkellernyc', 'biblegateway', 'SheReadsTruth', 'CSCleve', 'JohnFea1', ' DruHart', 'ayjay', ' TheAlanNoble', 'james_ka_smith', 'alissamarie', 'Fred_Buechner', 'ahc', ' rachelheldevans', 'mafeinberg', 'jonestony', 'scotmcknight', 'CarolHoward', 'JamesMartinSJ', 'TweetEnuma', 'EugeneCho', 'josaxton', 'ChristineCaine', 'drmoore', 'JemarTisby', 'BethMooreLPM', 'DerwinLGray', 'richvillodas', 'louiegiglio', 'esaumccaulley','hamzayusuf', 'mohamedghilan', 'tariqramadan', 'kzmawrid', 'JavedGhamidi', 'musafurber', 'ShaykhNinowy', 'MoAnsar', 'FarazRabbani', 'alhabibali', 'TahirulQadri', 'JihadiJew', 'TahirAshrafi', 'MuhammadBadsha', 'TCNLive', 'onfaith', 'Qantara', 'channelislam','Pontifex','DalaiLama','rickwarren','TariqRamadan','yusufchambers','JoyceMeyer','JoelOsteen','RaviZacharias','EckhartTolle','SadhguruJ','imamomarsuleiman','BishopJakes','HinduAmerican','MuslimAdvocates','CardinalChito','SisterRuthEO','RabbiWolpe','IngridMattson','RevJacquiLewis','JonathanACBrown','bobrobertsjr','Jentezen','muftimenk','JohnPiper','alainemerson','CardinalJWTobin','preetbharara','frankmschaefer','FatherRosica','SherlySantosJoy','ajcoss','LaryciaHawkins','christomlin','BethMooreLP','PadreSJ','BrianZahnd','alislam','RaviKZacharia','MiroslavVol','IyanlaVanza','BishopBarron','ImamSuhaibWebb','RevJacquiLewis','pastorbrady','LigonDuncan','LeilaAhmed','AminaWadud','omarsuleiman504','FrEdwardBeck','SakyongMipham'])),
        'racism': list(dict.fromkeys(['TheRevAl','Kaepernick7','deray','MsPackyett','ProfessorCrunk','nhannahjones','AfricanAmerNews','ShaunKing','Nettaaaaaaaa','opalayo','sandylocks','BreeNewsome','Colorlines','michele_norris','TaNehisiCoats','jelani9','TheJuanWilliams','keithboykin','repjohnlewis','angela_rye','ClintSmithIII','TiffanyDCross','Eugene_Scott','AOC','Maria_Hinojosa','Baratunde','Bakari_Sellers','CharlesMBlow','nhbaptiste','ava','deraysvest','repmaxinewaters','deray','angryblacklady','TalibKweli','kevin_powell','RashidaTlaib','CornelWest','repblumenauer','VanJones68','SybrinaFulton','lizzo','repjohnlewis','DerrickNAACP','K_JeanPierre','donlemon','thelasentinel','AprilDRyan','NCRMuseum','ColorOfChange','Bree Newsome', 'MsKellyMHayes', 'nowhitesaviors', 'strongblacklead', 'survivepunish', 'Blklivesmatter'])),
        'gender': list(dict.fromkeys(['Malala','EmmaWatson','UN_Women','HeForShe','ChimamandaReal','LetToysBeToys','melindagates','TaranaBurke','sherylsandberg','WomenintheWorld','iHollaback','HeforSheGhana','phumzileunwomen','MalalaFund','womensmarch','CecileRichards','Alyssa_Milano','Caitlyn_Jenner','MichelleObama','UNWomenAust','PlanGlobal','EverydaySexism','RAINN','TheEllenShow','thandienewton','UNYouthEnvoy','HillaryClinton','schemaly','UN_CSW','GirlUp','Trevornoah','UN_WomenWatch','YaraShahidi','priyankachopra','JessWeiner','frankiegrad','AfrozShah1','brielarson','TheDesmondTutu','TheRitaMoreno','UN_WomenUK','HelenClarkNZ','MalikaSaadaSaar','raindovemodel','GloriaSteinem','Lavernecox','TeenVogue','camanpour','TheGirlGen','Padmasree','TheAdvocateMag', 'glaad', 'outmagazine', 'LambdaLegal', 'huffpostqueer', 'Lavernecox', 'PinkNews', 'maddow', 'Queerty', 'gaystarnews', 'tlrd', 'stonewalluk'])),
        'politics': list(dict.fromkeys(['BarackObama','HillaryClinton','elonmusk','AOC','narendramodi','BorisJohnson','JustinTrudeau','tedcruz','KamalaHarris','RandPaul','POTUS','jeremycorbyn','nayibbukele','jacindaardern','EmmanuelMacron','SenSanders','JoeBiden','DNC','GOP','SenWarren','SpeakerPelosi','mehdirhasan','chrislhayes','andersoncooper','BBCPolitics','FoxNewsPolitics','seanhannity','TuckerCarlson','jaketapper','maddow','johnlegend','AprilDRyan','AriMelber','donnabrazile','FareedZakaria','Trevornoah','davidaxelrod','ananavarro','mitchellreports','HardballChris','megynkelly','seanspicer','GStephanopoulos','CoryBooker','LindseyGrahamSC','marcorubio','tedlieu','RepMaxineWaters','TulsiGabbard','TuckerCarlson','nprpolitics', 'McClatchyDC', 'SwingState', 'Wonkette', 'GOP12', 'senatus', 'politifact', 'CookPolitical'])),
        'disability': list(dict.fromkeys(['debraruh','JHMarble','robbyneff','WeAreAbleHere','SmartAssCripple','RuthMadeley','DisabilityIN','SpecialOlympics','BeckyMotivates','axschat','AndyAUCD','Ability2thrive','jennylayfluffy','deanofhealth','lizditz','sbaroncohen','Andrew_Pulrang','CDisability','crippledscholar','Easterseals','EnableMagazine','finnigr','LaurenAppelbaum','heidilgardner','JessSmith27','kenthehr','kentbuse','KulikovUNIATF','LFLegal','maysoonzayid','michele_jokinen','RealEvaK','rossetti_p','RuhGlobal','stephanieannd','SteveBrownGBWR','TomHanks','vj44','WalkandRollPEI','WDSDConference','AxelLeblois','AccessibleMedia','cospaad','DisRightsUK','EnableIndia','johnmuldoon','RNIB','Scope','UsherCoalition','wheelchairdancr','SinsInvalid', 'VilissaThompson', 'Sblahov', 'behearddc', 'GreggBeratan', 'Keah_Maria', 'DawnMGibson', 'autselfadvocacy', 'DiorVargas', 'DisabledLatinx', 'CWW', 'yovimi']))  
    }
    hashtags_dict = {
        'religion': list(dict.fromkeys([
            '#Religion','#Interfaith','#InterfaithDialogue','#Spirituality','#Faith','#ReligiousFreedom','#Interreligious','#InterfaithUnity','#InterfaithHarmony','#InterfaithUnderstanding','#InterfaithCommunity','#ReligiousTolerance','#InterfaithPeace','#InterfaithLove','#InterfaithCooperation','#ReligiousPluralism','#InterfaithRelations','#InterfaithLeadership','#FaithInAction','#InterfaithBridge','#InterfaithInclusion','#ReligiousDiversity','#InterfaithSolidarity','#FaithAndBelief','#InterfaithUnderstanding','#InterfaithEngagement','#ReligionAndCulture','#InterfaithRespect','#InterfaithEducation','#InterfaithDialogueWeek','#ReligionInSociety','#FaithBased','#InterfaithYouth','#InterfaithPrayer','#ReligiousHarmony','#InterfaithWeek','#InterfaithJustice','#FaithCommunity','#ReligiousEquality','#InterfaithLeaders','#InterfaithAdvocacy','#InterfaithUnderstanding','#ReligionAndPolitics','#InterfaithService','#InterfaithUnderstanding','#ReligionToday','#InterfaithCohesion','#InterfaithNetworking','#FaithInHumanity','#ReligiousHarmony'
        ])),
        'racism': list(dict.fromkeys([
            '#Racism','#EndRacism','#AntiRacism','#BlackLivesMatter','#Equality','#JusticeForAll','#SocialJustice','#NoHate','#RacialJustice','#StandAgainstRacism','#Inclusion','#EqualityForAll','#SayNoToRacism','#DiversityandInclusion','#HumanRights','#StopHate','#FightRacism','#TogetherAgainstRacism','#UnitedAgainstRacism','#Fairness','#RacialEquality','#ColorBlind','#UnityInDiversity','#CivilRights','#LoveNotHate','#WeAreOne','#EqualOpportunity','#BreakTheCycle','#UnlearnRacism','#ListenAndLearn','#EducateYourself','#Notoracism','#SayTheirNames','#StampedOutRacism','#RaceRelations','#SystemicRacism','#EndDiscrimination','#Solidarity','#KnowYourHistory','#Liberation'
        ])),
        'gender': list(dict.fromkeys([
            '#GenderEquality','#Feminism','#WomensRights','#HeForShe','#WomenEmpowerment','#GenderJustice','#EqualityForAll','#GenderInclusion','#BreakingTheGlassCeiling','#GenderEquity','#LGBTQ+','#GenderDiversity','#InclusiveLeadership','#EqualOpportunity','#TransRights','#DiverseVoices','#IntersectionalFeminism','#GenderRoles','#EmpowerWomen','#MenForEquality','#GenderStereotypes','#NonBinary','#GirlsInSTEM','#WomenInLeadership','#FeministFriday','#WomensHistoryMonth','#TransEquality','#GenderFluid','#QueerRights','#GenderBalance','#FeministMovement','#EqualPay','#GenderAwareness','#ChampionEquality','#LoveIsLove','#GirlsEducation','#ShePersisted','#GenderActivism','#DiversityMatters','#GenderHarmony','#TransVisibility','#FeministVoices','#AllGenderRestroom','#GenderRevolution','#EradicateSexism','#EmpowerGirls','#MenSupportingWomen','#DismantleThePatriarchy','#RespectHerVoice','#BeyondBinary'
        ])),
        'politics': list(dict.fromkeys([
            '#Politics','#PoliticalNews','#CurrentAffairs','#Democracy','#Government','#Policy','#Election','#Vote','#PoliticalScience','#Campaign2024','#PoliticalDebate','#PublicPolicy','#PoliticalAnalysis','#WorldPolitics','#GlobalAffairs','#CampaignTrail','#Vote2024','#PolicyMaking','#PoliticalLeadership','#PoliticalReform','#CivilRights','#InternationalRelations','#PoliticalIssues','#VoteResponsibly','#PoliticalActivism','#GovernmentWatch','#PoliticalDiscourse','#PoliticalCommentary','#StateOfTheUnion','#PowerToThePeople','#Governance','#ElectionCoverage','#CampaignTrail2024','#PoliticalInsider','#PoliticalThought','#CivicEngagement','#PoliticalParties','#VoteSmart','#PolicyDebate','#LeadershipMatters','#PoliticalCulture','#CampaignSeason','#PoliticalDialogue','#Decision2024','#PoliticalLeaders','#VoteYourVoice','#GovernmentPolicy','#PoliticalForum','#ElectionDay','#VoteDemocracy'
        ])),
        'disability': list(dict.fromkeys([
            '#Disability','#InclusionMatters','#Accessibility','#DisabilityRights','#InclusiveDesign','#AccessibleWorld','#DisabledAndProud','#Ableism','#DisabilityAwareness','#AccessibleTech','#InclusiveEducation','#DisabilityCommunity','#DisabilityAdvocacy','#InclusiveEmployment','#DisabilityJustice','#AssistiveTechnology','#DisabledLife','#AccessibilityForAll','#InclusionRevolution','#AbilityNotDisability','#SpecialNeeds','#PWD (Persons With Disabilities)','#InclusiveCulture','#AccessForAll','#DiverseAbilities','#EqualOpportunity','#DisabilitySupport','#AccessibleTravel','#DisabilityInclusion','#WheelchairLife','#AutismAwareness','#DisabilityRightsAreHumanRights','#ChronicIllness','#InclusiveSpaces','#DeafCommunity','#DisabilityEmployment','#Neurodiversity','#BlindAwareness','#AdaptiveSports','#DisabilityEquality','#DisabledVoices','#VisibleDisabilities','#MentalHealthAwareness','#DisabilityStudies','#AccessibleHousing','#InclusiveTechnology','#DisabilityPride','#InclusiveLeadership','#AccessibleTransportation','#CripTheVote'
        ]))  
    }
    dates_dict = get_dates_dict(datetime.date(2022, 1, 1), datetime.date.today(), 12)
    
    for key in list(words_dict.keys()):
        print(f'Category: {key}\nNumber of related accounts - {len(accaunts_dict[key])}, hashtags - {len(hashtags_dict[key])}, words - {len(words_dict[key])}')
    # options = {
    #     'categories': list(words_dict.keys()),
    #     'hashtags': hashtags_dict,
    #     'users': accaunts_dict,
    #     'words': words_dict,
    #     'dates': dates_dict
    # }

    # options = {
    #     'categories': ['racism', 'politics', 'gender', 'disability'],
    #     'hashtags': {},
    #     'users': accaunts_dict,
    #     'words': words_dict,
    #     'dates': dates_dict
    # }

    # queries = create_query_queue(options)
 
    # queue = asyncio.Queue()

    # workers_count = 18  # limit concurrency here 2 concurrent requests at time
    # workers = [asyncio.create_task(worker(queue, api)) for _ in range(workers_count)]

    # for q in queries:
    #     queue.put_nowait(q)

    # await queue.join()

    # for worker_task in workers:
    #     worker_task.cancel()

    # if queue.qsize() == 0:
    #     exit(0)    

if __name__ == "__main__":
    asyncio.run(main())